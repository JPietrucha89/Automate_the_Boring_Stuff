import openpyxl
import os
from openpyxl.styles import Font  # convenient for messing with fonts
from pathlib import Path

wb = openpyxl.Workbook()
sheet = wb[wb.sheetnames[0]]

# FONTS --------------------------------------------------------
# Keyword argument      Data type   Description
# name                  String      The font name, such as 'Calibri' or 'Times New Roman'
# size                  Integer     The point size
# bold                  Boolean     True, for bold font
# italic                Boolean     True, for italic font

# create Font object and then assign it to cell's attribute .font
fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

# FORMULAS --------------------------------------------------------
for i in range(4, 8):
    sheet.cell(i, 1).value = i
sheet['A8'] = '=SUM(A4:A7)'  # set formula

# ROW HEIGHT AND COLUMN WIDTH --------------------------------------------------------
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20

# MERGING AND UNMERGING CELLS --------------------------------------------------------
sheet.merge_cells('A1:D3')  # Merge all these cells.

# FREEZE PANES --------------------------------------------------------
# Note that all rows above and all columns to the left of this cell will be frozen, but the row and column of the cell itself will not be frozen.
sheet.freeze_panes = 'A2'  # Freeze the rows above A2.
# to unfreeze
sheet.freeze_panes = 'A1'
sheet.freeze_panes = None

wb.save(Path(os.getcwd(), 'working_with_Excel', 'excel_tinkering.xlsx'))
wb.close

# CHART --------------------------------------------------------
# OpenPyXL supports creating bar, line, scatter, and pie charts using the data in a sheet’s cells. To make a chart, you need to do the following:

# Create a Reference object from a rectangular selection of cells.
# Create a Series object by passing in the Reference object.
# Create a Chart object.
# Append the Series object to the Chart object.
# Add the Chart object to the Worksheet object, optionally specifying which cell should be the top-left corner of the chart.

# The Reference object requires some explaining. You create Reference objects by calling the openpyxl.chart.Reference() function and passing three arguments:
# The Worksheet object containing your chart data.
# A tuple of two integers, representing the top-left cell of the rectangular selection of cells containing your chart data: the first integer in the tuple is the row, and the second is the column. Note that 1 is the first row, not 0.
# A tuple of two integers, representing the bottom-right cell of the rectangular selection of cells containing your chart data: the first integer in the tuple is the row, and the second is the column.
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):  # create some data in column A
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(
    sheet, min_col=1, min_row=1, max_col=1, max_row=10)  # create Reference object
seriesObj = openpyxl.chart.Series(
    refObj, title='First series')  # create Series object

chartObj = openpyxl.chart.BarChart()  # create Chart object
chartObj.title = 'My Chart'
chartObj.append(seriesObj)  # append Series object to Chart object

sheet.add_chart(chartObj, 'C5')  # add Chart object to given location/address
wb.save(Path(os.getcwd(), 'working_with_Excel', 'sampleChart.xlsx'))
wb.close
