import openpyxl
import os
from pathlib import Path
from openpyxl.styles import Font


def create_wb_with_multiplication_table(table_size):
    wb = openpyxl.Workbook()
    sheet = wb[wb.sheetnames[0]]
    fontObj1 = Font(bold=True)
    for i in range(1, table_size+1):
        sheet.cell(1, 1+i).value = i  # create first row
        sheet.cell(1, 1+i).font = fontObj1  # bold=True
        sheet.cell(1+i, 1).value = i  # create first column
        sheet.cell(1+i, 1).font = fontObj1  # bold=True

    for i in range(1, table_size+1):
        for j in range(1, table_size+1):
            sheet.cell(1+i, 1+j).value = i * j

    wb.save(Path(os.getcwd(), 'working_with_Excel', 'table.xlsx'))
    wb.close


# main
create_wb_with_multiplication_table(6)
