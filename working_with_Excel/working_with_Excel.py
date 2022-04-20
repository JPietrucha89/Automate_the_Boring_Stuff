import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string

file=Path(Path.cwd(),'working_with_Excel','example.xlsx')
# open individual excel file and create workbook object
wb=openpyxl.load_workbook(file)

# get names of sheets and values of cells
print(wb.sheetnames)
sheet=wb['Sheet1']
print(type(sheet))

print(sheet.max_row) # Get the highest row number
print(sheet.max_column) # Get the highest column number

print(sheet['A1'].value)
print(str(sheet['A1'].value))
print(sheet['B1'].value)
print(sheet['C1'].value)

cell=sheet['B1']
print('Row is ' + str(cell.row) + ', column is ' + str(cell.column) + ', address is ' + cell.coordinate)
print(f'Row is {cell.row}, column is {cell.column}, address is {cell.coordinate}')
print(get_column_letter(900)) # return letter of column with given index
print(column_index_from_string('AA')) # return index of column with given letter

activeSheet = wb.active # Get the activeSheet
print(activeSheet.title)

# using .cell() method allows to pass columns name as number - useful for loops
print(sheet.cell(5,2).value)
for i in range(1,8):
    print(i,sheet.cell(i,2).value)

# create range and loop through all cells within 
range=sheet['A1':'C3']
for rowOfCellObjects in range:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

# To access the values of cells in a particular row or column, you can also use a Worksheet object’s rows and columns attribute. These attributes must be converted to lists with the list() function before you can use the square brackets and an index with them.
print(list(sheet.columns)[1]) # Get second column's cells.
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)

# create new worbook in memory
wb=openpyxl.Workbook()
print(wb.sheetnames) #default name of first sheet is Sheet
sheet=wb['Sheet']
sheet['A1'].value=10 # changing a cell's value is done using square brackets, just like changing values of list or dictionary
sheet['A2'].value='Python'
wb.save(Path(Path.cwd(),'working_with_Excel','new_workbook.xlsx'))

sheet2=wb.create_sheet() # add sheet with deafult name at the end of 
print(wb.sheetnames)
sheet2.title='My new sheet name' # rename sheet
print(wb.sheetnames)
wb.save(Path(Path.cwd(),'working_with_Excel','new_workbook_with_2_sheets.xlsx'))

sheet3=wb.create_sheet(index=0,title='My newest sheet name') # index allows to pick position of new worksheet - just like before/after:= in VBA
print(wb.sheetnames)
wb.save(Path(Path.cwd(),'working_with_Excel','new_workbook_with_3_sheets.xlsx'))

#deleting sheet
sheet4=wb.create_sheet()
del wb[sheet4.title]