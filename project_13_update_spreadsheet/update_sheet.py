import openpyxl, os
from pathlib import Path 

dict_price={'Celery' : 1.19,'Garlic' : 3.07,'Lemon' : 1.27}

p=Path(os.getcwd(),'project_13_update_spreadsheet','produceSales.xlsx')
wb=openpyxl.load_workbook(p)
print(wb.sheetnames)
sheet=wb[wb.sheetnames[0]]
number_of_rows=sheet.max_row

for i in range(2,number_of_rows+1):
    produce=sheet.cell(i,1).value
    price=sheet.cell(i,2).value
    if produce in dict_price:
        sheet.cell(i,2).value=dict_price[produce]

# save and close excel file
wb.save(Path(os.getcwd(),'project_13_update_spreadsheet','produceSales_corrected.xlsx'))
wb.close