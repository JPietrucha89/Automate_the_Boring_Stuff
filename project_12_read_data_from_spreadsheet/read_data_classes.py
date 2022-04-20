# In this project, you’ll write a script that can read from the census spreadsheet file and calculate statistics for each county in a matter of seconds.

# define Class
class County:
    def __init__(self,county_name,state,population) -> None:
        self.pop=population
        self.num_tracts=1
        self.county_name=county_name
        self.state=state

import openpyxl, pprint
from pathlib import Path

# Reads the data from the Excel spreadsheet
filename=Path(Path.cwd(),'project_12_read_data_from_spreadsheet','censuspopdata.xlsx')
wb=openpyxl.open(filename)
print(wb.sheetnames)
sheet=wb[wb.sheetnames[0]]

# create dictionary which will store class County objects
dict_county={}

for row in range(2,sheet.max_row+1):
    state  = sheet['B' + str(row)].value
    county_name = sheet['C' + str(row)].value
    pop    = sheet['D' + str(row)].value
    if county_name not in dict_county:
        new_county=County(county_name,state,pop)
        dict_county[county_name]=new_county
    else:
        dict_county[county_name].pop+=pop
        dict_county[county_name].num_tracts+=1

# Print the results
# Open a new excel file and write the contents of dict_county to it.
print('Writing results to *.xlsx file...')
wbResult=openpyxl.Workbook()

sheet=wbResult[wbResult.sheetnames[0]]
sheet['A1'].value='County'
sheet['B1'].value='State'
sheet['C1'].value='Total population'
sheet['D1'].value='Number of tracts'

sheet_counter=2
for key in dict_county.keys():
    sheet.cell(sheet_counter,1).value=key
    sheet.cell(sheet_counter,2).value=dict_county[key].state
    sheet.cell(sheet_counter,3).value=dict_county[key].pop
    sheet.cell(sheet_counter,4).value=dict_county[key].num_tracts
    sheet_counter+=1

print('Done.')
# write whole dict_county to new python file which can be imported and used later -> read_from_census.py
wbResult.save(Path(Path.cwd(),'project_12_read_data_from_spreadsheet','census2010_classes.xlsx'))
wbResult.close()
wb.close()