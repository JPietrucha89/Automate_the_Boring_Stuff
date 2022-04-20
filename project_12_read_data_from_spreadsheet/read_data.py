# In this project, you’ll write a script that can read from the census spreadsheet file and calculate statistics for each county in a matter of seconds.

# This is what your program does:
# Reads the data from the Excel spreadsheet
# Counts the number of census tracts in each county
# Counts the total population of each county
# Prints the results

# This means your code will need to do the following:
# Open and read the cells of an Excel document with the openpyxl module.
# Calculate all the tract and population data and store it in a data structure.
# Write the data structure to a text file with the .py extension using the pprint module.

import openpyxl, pprint
from pathlib import Path

# Reads the data from the Excel spreadsheet
filename=Path(Path.cwd(),'project_12_read_data_from_spreadsheet','censuspopdata.xlsx')
wb=openpyxl.open(filename)
print(wb.sheetnames)
sheet=wb[wb.sheetnames[0]]

# create empty dictionary, with state as keys, and subdictionaries as values
# subdictionary will have county as keys, and subsubdictionaries as values
# subsubdictionary will have populations/num_tracts as keys and integers as values
# structure [state][county][population/tracts] will be added later:
# {'AK': {'Aleutians East': {'pop': 3141, 'tracts': 1},
#         'Aleutians West': {'pop': 5561, 'tracts': 2},
#         'Anchorage': {'pop': 291826, 'tracts': 55},
#         'Bethel': {'pop': 17013, 'tracts': 3},
#         'Bristol Bay': {'pop': 997, 'tracts': 1},
#         --snip--
dict_county={}

for row in range(2,sheet.max_row+1):
    state  = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop    = sheet['D' + str(row)].value

    # Make sure the key for this state exists. If it exists nothing happens
    dict_county.setdefault(state, {})
    # Make sure the key for this county in this state exists. If it exists nothing happens
    dict_county[state].setdefault(county, {'num_tracts': 0, 'pop': 0})

    # Count the number of census tracts in each county
    # Each row represents one census tract, so increment by one.
    dict_county[state][county]['num_tracts'] += 1

    # Count the total population of each county
    # Increase the county pop by the pop in this census tract.
    dict_county[state][county]['pop'] += int(pop)

# Print the results
# Open a new text file and write the contents of countyData to it.
print('Writing results...')
resultFile = open(Path(Path.cwd(),'project_12_read_data_from_spreadsheet','census2010.py'), 'w')
resultFile.write('allData = ' + pprint.pformat(dict_county)) # write whole dict_county to new python file which can be imported and used later -> read_from_census.py
resultFile.close()
print('Done.')

wb.close()